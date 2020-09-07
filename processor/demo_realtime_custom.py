#!/usr/bin/env python
import os
import sys
import argparse
import json
import shutil
import time

import numpy as np
import torch
import skvideo.io

from .io import IO
import tools
import tools.utils as utils

import cv2

from openpyxl import Workbook

from lightOpenPose.lightOpenPose import LightOpenPose

import threading
from tools.utils.WarningPrint import WarningPrint


class ipcamCapture:
    def __init__(self, URL):
        self.Frame = []
        self.status = False
        self.isstop = False
        # 攝影機連接。
        self.capture = cv2.VideoCapture(URL)

    def start(self):
        # 把程式放進子執行緒，daemon=True 表示該執行緒會隨著主執行緒關閉而關閉。
        print('ipcam started!')
        threading.Thread(target=self.queryframe, daemon=True, args=()).start()

    def stop(self):
        # 記得要設計停止無限迴圈的開關。
        self.isstop = True
        print('ipcam stopped!')

    def getframe(self):
        # 當有需要影像時，再回傳最新的影像。
        return self.Frame

    def queryframe(self):
        # 不斷讀取直到沒有更新的畫面
        while (not self.isstop):
            self.status, self.Frame = self.capture.read()
        self.capture.release()

class DemoRealtime(IO):
    """ A demo for utilizing st-gcn in the realtime action recognition.
    The Openpose python-api is required for this demo.

    Since the pre-trained model is trained on videos with 30fps,
    and Openpose is hard to achieve this high speed in the single GPU,
    if you want to predict actions by **camera** in realtime,
    either data interpolation or new pre-trained model
    is required.

    Pull requests are always welcome.
    """

    def start(self):
        # 自訂參數
        write_stgcn_video = True
        stgcn_imshow = False
        write_custom_video = True # True imshow
        write_excel = True
        useLightOpenpose = False

        if write_stgcn_video or write_custom_video:
            fps = 30
            fourcc = cv2.VideoWriter_fourcc(*'XVID')
            video_name = self.arg.video.split('/')[-1].split('.')[0]
            if write_stgcn_video:
                out = cv2.VideoWriter('data/mydata/test_output/' + video_name + '.avi', fourcc, fps, (1918, 1080))
            else:
                out = None
            if write_custom_video:
                out_custom_demo = cv2.VideoWriter('data/mydata/test_output_custom_demo/' + video_name + '.avi', fourcc, fps, (1920, 1080))
            else:
                out_custom_demo = None
        
        if write_excel:
            # 建立excel檔 (報表)
            wb = Workbook()
            sheet = wb.active
            excelSavePath = 'data/mydata/test_output_custom_excel/' + video_name + '.xlsx'

        warningPrint = WarningPrint()

        # load openpose python api
        if self.arg.openpose is not None:
            sys.path.append('{}/python'.format(self.arg.openpose))
            sys.path.append('{}/build/python'.format(self.arg.openpose))
        try:
            from openpose import pyopenpose as op
        except:
            print('Can not find Openpose Python API.')
            return

        label_name_path = './resource/custom_dataset/label_name.txt'
        with open(label_name_path) as f:
            label_name = f.readlines()
            label_name = [line.rstrip() for line in label_name]
            self.label_name = label_name

        # initiate openpose
        if useLightOpenpose:
            lightOpenPose = LightOpenPose(256)
        else:
            opWrapper = op.WrapperPython()
            params = dict(model_folder='./models', model_pose='COCO')
            opWrapper.configure(params)
            opWrapper.start()
        # initiate
        self.model.eval()
        pose_tracker = naive_pose_tracker()

        if self.arg.video == 'camera_source':
            video_capture = cv2.VideoCapture(0)
        elif self.arg.video == 'ip_camera':
            video_path = 'rtsp://admin:admin@192.168.1.245:554/channel1'
            # video_capture = cv2.VideoCapture(video_path)
            ipcam = ipcamCapture(video_path)
            ipcam.start()
            time.sleep(1)
        else:
            video_capture = cv2.VideoCapture(self.arg.video)

        # start recognition
        start_time = time.time()
        frame_index = 0


        while(True):

            tic = time.time()

            # get image
            if self.arg.video == 'ip_camera':
                orig_image = ipcam.getframe()
            else:
                ret, orig_image = video_capture.read()
            if orig_image is None:
                break
            if write_custom_video:
                orig_image_for_show = orig_image
            else:
                orig_image_for_show = None
            source_H, source_W, _ = orig_image.shape
            orig_image = cv2.resize(
                orig_image, (256 * source_W // source_H, 256))
            H, W, _ = orig_image.shape
            
            time_after_get_img = time.time()
            # print('get_img: ' + str(time_after_get_img - tic))

            # pose estimation
            if useLightOpenpose:
                multi_pose = lightOpenPose.getPose(orig_image)
                if multi_pose.shape[0] == 0:
                    continue
            else:
                datum = op.Datum()
                datum.cvInputData = orig_image
                opWrapper.emplaceAndPop([datum])
                multi_pose = datum.poseKeypoints  # (num_person, num_joint, 3)

            if len(multi_pose.shape) != 3:
                continue

            # normalization
            multi_pose[:, :, 0] = multi_pose[:, :, 0]/W
            multi_pose[:, :, 1] = multi_pose[:, :, 1]/H
            multi_pose[:, :, 0:2] = multi_pose[:, :, 0:2] - 0.5
            multi_pose[:, :, 0][multi_pose[:, :, 2] == 0] = 0
            multi_pose[:, :, 1][multi_pose[:, :, 2] == 0] = 0

            time_get_pose = time.time()
            # print('get pose: ' + str(time_get_pose - time_after_get_img))

            # pose tracking
            if self.arg.video == 'camera_source' or self.arg.video == 'ip_camera':
                frame_index = int((time.time() - start_time)*self.arg.camera_fps)
            else:
                frame_index += 1
            pose_tracker.update(multi_pose, frame_index)
            data_numpy = pose_tracker.get_skeleton_sequence()
            data = torch.from_numpy(data_numpy)
            data = data.unsqueeze(0)
            data = data.float().to(self.dev).detach()  # (1, channel, frame, joint, person)

            time_pose_tracking = time.time()
            # print('pose_tracking: ' + str(time_pose_tracking - time_get_pose))

            # model predict
            # if frame_index % 15 == 0 or frame_index == 1:
            voting_label_name, video_label_name, output, intensity = self.predict(
                data)

            time_after_predict = time.time()
            # print('predict: ' + str(time_after_predict - time_pose_tracking))

            # visualization
            if stgcn_imshow or write_stgcn_video:
                skip_stgcn_img = False
            else:
                skip_stgcn_img = True
            app_fps = 1 / (time.time() - tic)
            image = self.render(data_numpy, voting_label_name,
                                video_label_name, intensity, orig_image, orig_image_for_show, out_custom_demo, frame_index, sheet, skip_stgcn_img, warningPrint, app_fps)
            if stgcn_imshow:
                cv2.imshow("ST-GCN", image)
            if write_stgcn_video:
                # image = cv2.resize(image, (1918, 1080))
                out.write(image)

            time_after_show_result = time.time()
            # print('show result: ' + str(time_after_show_result - time_after_predict))

            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

            # print('all time: ' + str(time.time() - tic))

        totle_time = time.time() -start_time
        # print('totle time: ' + str(totle_time))

        if write_excel:
            for col in range(sheet.max_column - 1):
                insertCell = sheet.cell(row=1, column=col + 2)
                insertCell.value = 'ID: {}'.format(col + 1)
            # 在報表左側每一列標上時間，方便觀看
            for i in range(0, (frame_index // 15) + 1):
                insertCell = sheet.cell(row=i + 2, column=1)
                insertCell.value = '{:.1f}秒'.format((i * 0.5))
            wb.save(excelSavePath)

    def predict(self, data):
        # forward
        output, feature = self.model.extract_feature(data)
        output = output[0]
        # print(output)
        feature = feature[0]
        intensity = (feature*feature).sum(dim=0)**0.5
        intensity = intensity.cpu().detach().numpy()

        # get result
        # classification result of the full sequence
        voting_label = output.sum(dim=3).sum(
            dim=2).sum(dim=1).argmax(dim=0)
        voting_label_name = self.label_name[voting_label]
        # classification result for each person of the latest frame
        num_person = data.size(4)
        latest_frame_label = [output[:, :, :, m].sum(
            dim=2)[:, -1].argmax(dim=0) for m in range(num_person)]
        latest_frame_label_name = [self.label_name[l]
                                   for l in latest_frame_label]

        num_person = output.size(3)
        num_frame = output.size(1)
        video_label_name = list()
        for t in range(num_frame):
            frame_label_name = list()
            for m in range(num_person):
                person_label = output[:, t, :, m].sum(dim=1).argmax(dim=0)
                person_label_name = self.label_name[person_label]
                frame_label_name.append(person_label_name)
            video_label_name.append(frame_label_name)
        return voting_label_name, video_label_name, output, intensity

    def render(self, data_numpy, voting_label_name, video_label_name, intensity, orig_image, orig_image_for_show, out_custom_demo, frame_index, sheet, skip_stgcn_img, warningPrint, fps=0):
        label_sequence = video_label_name[len(video_label_name) - 4:]
        images = utils.visualization.stgcn_visualize(
            data_numpy[:, [-1]],
            self.model.graph.edge,
            intensity[[-1]], [orig_image],
            orig_image_for_show,
            out_custom_demo,
            frame_index,
            sheet,
            skip_stgcn_img,
            warningPrint,
            voting_label_name,
            label_sequence,
            self.arg.height,
            fps=fps)
        if skip_stgcn_img:
            image = next(images)
            return image
        else:
            image = next(images)
            image = image.astype(np.uint8)
            return image

    @staticmethod
    def get_parser(add_help=False):

        # parameter priority: command line > config > default
        parent_parser = IO.get_parser(add_help=False)
        parser = argparse.ArgumentParser(
            add_help=add_help,
            parents=[parent_parser],
            description='Demo for Spatial Temporal Graph Convolution Network')

        # region arguments yapf: disable
        parser.add_argument('--video',
                            default='./resource/media/skateboarding.mp4',
                            help='Path to video')
        parser.add_argument('--openpose',
                            default=None,
                            help='Path to openpose')
        parser.add_argument('--model_input_frame',
                            default=128,
                            type=int)
        parser.add_argument('--model_fps',
                            default=30,
                            type=int)
        parser.add_argument('--camera_fps',
                            default=30,
                            type=int)
        parser.add_argument('--height',
                            default=1080,
                            type=int,
                            help='height of frame in the output video.')
        parser.set_defaults(
            config='./config/st_gcn/custom-skeleton/demo_realtime.yaml')
        parser.set_defaults(print_log=False)
        # endregion yapf: enable

        return parser

class naive_pose_tracker():
    """ A simple tracker for recording person poses and generating skeleton sequences.
    For actual occasion, I recommend you to implement a robuster tracker.
    Pull-requests are welcomed.
    """

    def __init__(self, data_frame=128, num_joint=18, max_frame_dis=np.inf):
        self.data_frame = data_frame
        self.num_joint = num_joint
        self.max_frame_dis = max_frame_dis
        self.latest_frame = 0
        self.trace_info = list()

    def update(self, multi_pose, current_frame):
        # multi_pose.shape: (num_person, num_joint, 3)

        if current_frame <= self.latest_frame:
            return

        if len(multi_pose.shape) != 3:
            return

        # score_order = (-multi_pose[:, :, 2].sum(axis=1)).argsort(axis=0)
        # print(score_order)
        # for p in multi_pose[score_order]:
        for p in multi_pose:

            # match existing traces
            matching_trace = None
            matching_dis = None
            for trace_index, (trace, latest_frame) in enumerate(self.trace_info):
                # trace.shape: (num_frame, num_joint, 3)
                if current_frame <= latest_frame:
                    continue
                mean_dis, is_close = self.get_dis(trace, p)
                if is_close:
                    if matching_trace is None:
                        matching_trace = trace_index
                        matching_dis = mean_dis
                    elif matching_dis > mean_dis:
                        matching_trace = trace_index
                        matching_dis = mean_dis

            # update trace information
            if matching_trace is not None:
                trace, latest_frame = self.trace_info[matching_trace]

                # padding zero if the trace is fractured
                pad_mode = 'interp' if latest_frame == self.latest_frame else 'zero'
                pad = current_frame-latest_frame-1
                new_trace = self.cat_pose(trace, p, pad, pad_mode)
                self.trace_info[matching_trace] = (new_trace, current_frame)

            else:
                new_trace = np.array([p])
                self.trace_info.append((new_trace, current_frame))

        self.latest_frame = current_frame

    def get_skeleton_sequence(self):

        # remove old traces
        valid_trace_index = []
        for trace_index, (trace, latest_frame) in enumerate(self.trace_info):
            if self.latest_frame - latest_frame < self.data_frame:
                valid_trace_index.append(trace_index)
        self.trace_info = [self.trace_info[v] for v in valid_trace_index]

        num_trace = len(self.trace_info)
        if num_trace == 0:
            return None

        data = np.zeros((3, self.data_frame, self.num_joint, num_trace))
        for trace_index, (trace, latest_frame) in enumerate(self.trace_info):
            end = self.data_frame - (self.latest_frame - latest_frame)
            d = trace[-end:]
            beg = end - len(d)
            data[:, beg:end, :, trace_index] = d.transpose((2, 0, 1))

        return data

    # concatenate pose to a trace
    def cat_pose(self, trace, pose, pad, pad_mode):
        # trace.shape: (num_frame, num_joint, 3)
        num_joint = pose.shape[0]
        num_channel = pose.shape[1]
        if pad != 0:
            if pad_mode == 'zero':
                trace = np.concatenate(
                    (trace, np.zeros((pad, num_joint, 3))), 0)
            elif pad_mode == 'interp':
                last_pose = trace[-1]
                coeff = [(p+1)/(pad+1) for p in range(pad)]
                interp_pose = [(1-c)*last_pose + c*pose for c in coeff]
                trace = np.concatenate((trace, interp_pose), 0)
        new_trace = np.concatenate((trace, [pose]), 0)
        return new_trace

    # calculate the distance between a existing trace and the input pose

    def get_dis(self, trace, pose):
        last_pose_xy = trace[-1, :, 0:2]
        curr_pose_xy = pose[:, 0:2]

        mean_dis = ((((last_pose_xy - curr_pose_xy)**2).sum(1))**0.5).mean()
        wh = last_pose_xy.max(0) - last_pose_xy.min(0)
        scale = (wh[0] * wh[1]) ** 0.5 + 0.0001
        is_close = mean_dis < scale * self.max_frame_dis
        return mean_dis, is_close
